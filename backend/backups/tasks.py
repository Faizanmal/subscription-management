"""
Celery tasks for backups app
"""

from celery import shared_task
from django.utils import timezone
from django.conf import settings
import logging
import os
import json
import hashlib
import zipfile

logger = logging.getLogger('swm')


@shared_task(bind=True, max_retries=3)
def run_backup(self, backup_id):
    """Run a backup job"""
    from backups.models import Backup
    
    try:
        backup = Backup.objects.select_related('schedule').get(id=backup_id)
        
        backup.status = 'running'
        backup.started_at = timezone.now()
        backup.save()
        
        try:
            # Get organization
            org = backup.schedule.organization
            
            # Create backup directory
            backup_dir = os.path.join(settings.BACKUP_ROOT, str(org.id))
            os.makedirs(backup_dir, exist_ok=True)
            
            # Create backup filename
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            filename = f"backup_{org.slug}_{timestamp}.zip"
            filepath = os.path.join(backup_dir, filename)
            
            # Collect data
            data = collect_backup_data(org, backup.metadata or {})
            
            # Create zip file
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Add JSON data
                zf.writestr('data.json', json.dumps(data, indent=2, default=str))
                
                # Add metadata
                zf.writestr('metadata.json', json.dumps({
                    'organization': str(org.id),
                    'organization_name': org.name,
                    'created_at': timezone.now().isoformat(),
                    'version': '1.0'
                }, indent=2))
            
            # Calculate checksum
            with open(filepath, 'rb') as f:
                checksum = hashlib.sha256(f.read()).hexdigest()
            
            # Update backup record
            backup.status = 'completed'
            backup.completed_at = timezone.now()
            backup.file_path = filepath
            backup.file_size = os.path.getsize(filepath)
            backup.checksum = checksum
            backup.save()
            
            # Update schedule
            backup.schedule.last_run_at = timezone.now()
            backup.schedule.save()
            
            logger.info(f"Backup completed: {filepath}")
            
            # Cleanup old backups
            cleanup_old_backups.delay(str(backup.schedule.id))
            
        except Exception as e:
            backup.status = 'failed'
            backup.error_message = str(e)
            backup.completed_at = timezone.now()
            backup.save()
            raise
        
    except Backup.DoesNotExist:
        logger.error(f"Backup {backup_id} not found")
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        self.retry(countdown=300)


def collect_backup_data(organization, options):
    """Collect all organization data for backup"""
    from users.models import User, Department, Team, Role
    from services.models import (
        Subscription, SubscriptionUser, UsageMetrics,
        CostRecord, Recommendation, Alert, Workflow
    )
    from integrations.models import Integration
    
    data = {
        'users': list(User.objects.filter(organization=organization).values()),
        'departments': list(Department.objects.filter(organization=organization).values()),
        'teams': list(Team.objects.filter(organization=organization).values()),
        'roles': list(Role.objects.filter(organization=organization).values()),
        'subscriptions': list(Subscription.objects.filter(organization=organization).values()),
        'subscription_users': list(
            SubscriptionUser.objects.filter(
                subscription__organization=organization
            ).values()
        ),
        'recommendations': list(
            Recommendation.objects.filter(
                subscription__organization=organization
            ).values()
        ),
        'alerts': list(
            Alert.objects.filter(
                subscription__organization=organization
            ).values()
        ),
        'workflows': list(Workflow.objects.filter(organization=organization).values()),
        'integrations': list(
            Integration.objects.filter(organization=organization).values(
                'id', 'name', 'type', 'status', 'config', 'created_at'
                # Exclude credentials for security
            )
        )
    }
    
    # Optionally include usage data (can be large)
    if options.get('include_usage', True):
        # Limit to last 90 days
        threshold = timezone.now() - timezone.timedelta(days=90)
        
        data['usage_metrics'] = list(
            UsageMetrics.objects.filter(
                subscription__organization=organization,
                period_start__gte=threshold.date()
            ).values()
        )
        
        data['cost_records'] = list(
            CostRecord.objects.filter(
                subscription__organization=organization,
                period_start__gte=threshold.date()
            ).values()
        )
    
    return data


@shared_task
def cleanup_old_backups(schedule_id):
    """Clean up old backups based on retention policy"""
    from backups.models import BackupSchedule, Backup
    
    try:
        schedule = BackupSchedule.objects.get(id=schedule_id)
        
        if schedule.retention_days == 0:
            # Keep forever
            return
        
        threshold = timezone.now() - timezone.timedelta(days=schedule.retention_days)
        
        old_backups = Backup.objects.filter(
            schedule=schedule,
            created_at__lt=threshold,
            status='completed'
        )
        
        for backup in old_backups:
            # Delete file
            if backup.file_path and os.path.exists(backup.file_path):
                os.remove(backup.file_path)
            
            backup.delete()
        
        if old_backups.count() > 0:
            logger.info(f"Cleaned up {old_backups.count()} old backups")
        
    except BackupSchedule.DoesNotExist:
        logger.error(f"Backup schedule {schedule_id} not found")


@shared_task(bind=True, max_retries=3)
def restore_backup(self, backup_id, options=None):
    """Restore from a backup"""
    from backups.models import Backup
    
    options = options or {}
    
    try:
        backup = Backup.objects.get(id=backup_id)
        
        if backup.status != 'completed':
            raise ValueError("Cannot restore from incomplete backup")
        
        if not backup.file_path or not os.path.exists(backup.file_path):
            raise ValueError("Backup file not found")
        
        # Extract and parse backup
        with zipfile.ZipFile(backup.file_path, 'r') as zf:
            data = json.loads(zf.read('data.json'))
            metadata = json.loads(zf.read('metadata.json'))
        
        # Verify organization matches
        org = backup.schedule.organization
        if str(org.id) != metadata.get('organization'):
            raise ValueError("Backup organization mismatch")
        
        # Restore data
        restore_data(org, data, options)
        
        logger.info(f"Restore completed from backup {backup_id}")
        
        return {'success': True}
        
    except Backup.DoesNotExist:
        logger.error(f"Backup {backup_id} not found")
        return {'success': False, 'error': 'Backup not found'}
    except Exception as e:
        logger.error(f"Restore failed: {e}")
        self.retry(countdown=60)


def restore_data(organization, data, options):
    """Restore organization data from backup"""
    from users.models import Department, Team, Role
    from services.models import Subscription, Recommendation
    
    # Restore in order of dependencies
    restore_entities = options.get('entities', ['all'])
    
    if 'all' in restore_entities or 'departments' in restore_entities:
        for dept_data in data.get('departments', []):
            Department.objects.update_or_create(
                id=dept_data['id'],
                defaults={k: v for k, v in dept_data.items() if k != 'id'}
            )
    
    if 'all' in restore_entities or 'teams' in restore_entities:
        for team_data in data.get('teams', []):
            Team.objects.update_or_create(
                id=team_data['id'],
                defaults={k: v for k, v in team_data.items() if k != 'id'}
            )
    
    if 'all' in restore_entities or 'roles' in restore_entities:
        for role_data in data.get('roles', []):
            Role.objects.update_or_create(
                id=role_data['id'],
                defaults={k: v for k, v in role_data.items() if k != 'id'}
            )
    
    if 'all' in restore_entities or 'subscriptions' in restore_entities:
        for sub_data in data.get('subscriptions', []):
            Subscription.objects.update_or_create(
                id=sub_data['id'],
                defaults={k: v for k, v in sub_data.items() if k != 'id'}
            )
    
    if 'all' in restore_entities or 'recommendations' in restore_entities:
        for rec_data in data.get('recommendations', []):
            Recommendation.objects.update_or_create(
                id=rec_data['id'],
                defaults={k: v for k, v in rec_data.items() if k != 'id'}
            )


@shared_task(bind=True, max_retries=3)
def run_data_export(self, export_id):
    """Run a data export job"""
    from backups.models import DataExport
    
    try:
        export = DataExport.objects.get(id=export_id)
        
        export.status = 'processing'
        export.started_at = timezone.now()
        export.save()
        
        try:
            org = export.organization
            
            # Collect data based on data_types
            data = collect_export_data(org, export.data_types, export.filters)
            
            # Create export directory
            export_dir = os.path.join(settings.EXPORT_ROOT, str(org.id))
            os.makedirs(export_dir, exist_ok=True)
            
            # Generate file
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            
            if export.format == 'json':
                filename = f"export_{timestamp}.json"
                filepath = os.path.join(export_dir, filename)
                with open(filepath, 'w') as f:
                    json.dump(data, f, indent=2, default=str)
            
            elif export.format == 'csv':
                filename = f"export_{timestamp}.zip"
                filepath = os.path.join(export_dir, filename)
                create_csv_export(data, filepath)
            
            elif export.format == 'xlsx':
                filename = f"export_{timestamp}.xlsx"
                filepath = os.path.join(export_dir, filename)
                create_excel_export(data, filepath)
            
            else:
                raise ValueError(f"Unsupported format: {export.format}")
            
            # Update export record
            export.status = 'completed'
            export.completed_at = timezone.now()
            export.file_path = filepath
            export.file_size = os.path.getsize(filepath)
            export.expires_at = timezone.now() + timezone.timedelta(days=7)
            export.save()
            
            # Notify user
            from users.models import Notification
            Notification.objects.create(
                user=export.requested_by,
                type='export',
                title='Data export ready',
                message='Your data export is ready for download.',
                action_url=f'/exports/{export.id}/download'
            )
            
            logger.info(f"Export completed: {filepath}")
            
        except Exception as e:
            export.status = 'failed'
            export.error_message = str(e)
            export.completed_at = timezone.now()
            export.save()
            raise
        
    except DataExport.DoesNotExist:
        logger.error(f"Export {export_id} not found")
    except Exception as e:
        logger.error(f"Export failed: {e}")
        self.retry(countdown=60)


def collect_export_data(organization, data_types, filters):
    """Collect data for export"""
    from services.models import Subscription, UsageMetrics, CostRecord, Recommendation, Alert
    from users.models import User, AuditLog
    
    data = {}
    
    # Parse date filters
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')
    
    if not data_types or 'subscriptions' in data_types:
        qs = Subscription.objects.filter(organization=organization)
        data['subscriptions'] = list(qs.values())
    
    if not data_types or 'usage' in data_types:
        qs = UsageMetrics.objects.filter(subscription__organization=organization)
        if date_from:
            qs = qs.filter(period_start__gte=date_from)
        if date_to:
            qs = qs.filter(period_end__lte=date_to)
        data['usage'] = list(qs.values())
    
    if not data_types or 'costs' in data_types:
        qs = CostRecord.objects.filter(subscription__organization=organization)
        if date_from:
            qs = qs.filter(period_start__gte=date_from)
        if date_to:
            qs = qs.filter(period_end__lte=date_to)
        data['costs'] = list(qs.values())
    
    if not data_types or 'recommendations' in data_types:
        qs = Recommendation.objects.filter(subscription__organization=organization)
        data['recommendations'] = list(qs.values())
    
    if not data_types or 'alerts' in data_types:
        qs = Alert.objects.filter(subscription__organization=organization)
        data['alerts'] = list(qs.values())
    
    if not data_types or 'users' in data_types:
        qs = User.objects.filter(organization=organization)
        data['users'] = list(qs.values('id', 'email', 'first_name', 'last_name', 'is_active', 'created_at'))
    
    if not data_types or 'audit_logs' in data_types:
        qs = AuditLog.objects.filter(organization=organization)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        data['audit_logs'] = list(qs.values())
    
    return data


def create_csv_export(data, filepath):
    """Create CSV export (zip of multiple CSVs)"""
    
    with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
        for key, records in data.items():
            if not records:
                continue
            
            # Create CSV in memory
            output = []
            headers = records[0].keys()
            output.append(','.join(headers))
            
            for record in records:
                row = [str(record.get(h, '')) for h in headers]
                output.append(','.join(f'"{v}"' for v in row))
            
            zf.writestr(f'{key}.csv', '\n'.join(output))


def create_excel_export(data, filepath):
    """Create Excel export"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        first_sheet = True
        for key, records in data.items():
            if not records:
                continue
            
            if first_sheet:
                ws = wb.active
                ws.title = key[:31]  # Excel sheet name limit
                first_sheet = False
            else:
                ws = wb.create_sheet(key[:31])
            
            # Add headers
            headers = list(records[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            for row_num, record in enumerate(records, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col, value=str(record.get(header, '')))
        
        wb.save(filepath)
        
    except ImportError:
        # Fall back to CSV if openpyxl not available
        create_csv_export(data, filepath.replace('.xlsx', '.zip'))


@shared_task(bind=True, max_retries=3)
def run_import_job(self, import_job_id):
    """Run an import job"""
    from backups.models import ImportJob
    
    try:
        job = ImportJob.objects.get(id=import_job_id)
        
        if job.status not in ['pending', 'validating']:
            return
        
        job.status = 'processing'
        job.started_at = timezone.now()
        job.save()
        
        try:
            temp_path = job.metadata.get('temp_path') if job.metadata else None
            
            if not temp_path or not os.path.exists(temp_path):
                raise ValueError("Import file not found")
            
            # Parse and import data
            result = process_import(job, temp_path)
            
            job.status = 'completed'
            job.completed_at = timezone.now()
            job.total_rows = result['total']
            job.processed_rows = result['processed']
            job.success_count = result['success']
            job.error_count = result['errors']
            job.errors = result.get('error_details', [])
            job.save()
            
            # Cleanup temp file
            os.remove(temp_path)
            
            # Notify user
            from users.models import Notification
            Notification.objects.create(
                user=job.uploaded_by,
                type='import',
                title='Import completed',
                message=f'Imported {result["success"]} of {result["total"]} records.',
                action_url=f'/imports/{job.id}'
            )
            
            logger.info(f"Import completed: {result}")
            
        except Exception as e:
            job.status = 'failed'
            job.error_message = str(e)
            job.completed_at = timezone.now()
            job.save()
            raise
        
    except ImportJob.DoesNotExist:
        logger.error(f"Import job {import_job_id} not found")
    except Exception as e:
        logger.error(f"Import failed: {e}")
        self.retry(countdown=60)


def process_import(job, filepath):
    """Process import file"""
    
    # Determine file type
    if filepath.endswith('.csv'):
        data = parse_csv(filepath)
    elif filepath.endswith('.json'):
        with open(filepath) as f:
            data = json.load(f)
    else:
        raise ValueError("Unsupported file format")
    
    result = {
        'total': len(data),
        'processed': 0,
        'success': 0,
        'errors': 0,
        'error_details': []
    }
    
    org = job.organization
    mapping = job.mapping or {}
    
    for i, row in enumerate(data):
        try:
            if job.import_type == 'subscriptions':
                import_subscription(org, row, mapping)
            elif job.import_type == 'users':
                import_user(org, row, mapping)
            elif job.import_type == 'usage':
                import_usage(org, row, mapping)
            else:
                raise ValueError(f"Unknown import type: {job.import_type}")
            
            result['success'] += 1
            
        except Exception as e:
            result['errors'] += 1
            result['error_details'].append({
                'row': i + 1,
                'error': str(e)
            })
        
        result['processed'] += 1
    
    return result


def parse_csv(filepath):
    """Parse CSV file"""
    import csv
    
    data = []
    with open(filepath, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    
    return data


def import_subscription(organization, row, mapping):
    """Import a subscription row"""
    from services.models import Subscription, SoftwareVendor
    
    # Apply mapping
    name = row.get(mapping.get('name', 'name'))
    if not name:
        raise ValueError("Name is required")
    
    vendor_name = row.get(mapping.get('vendor', 'vendor'))
    vendor = None
    if vendor_name:
        vendor, _ = SoftwareVendor.objects.get_or_create(name=vendor_name)
    
    Subscription.objects.update_or_create(
        organization=organization,
        name=name,
        defaults={
            'vendor': vendor,
            'status': row.get(mapping.get('status', 'status'), 'active'),
            'billing_cycle': row.get(mapping.get('billing_cycle', 'billing_cycle'), 'monthly'),
            'monthly_cost': float(row.get(mapping.get('monthly_cost', 'monthly_cost'), 0) or 0),
            'total_licenses': int(row.get(mapping.get('total_licenses', 'total_licenses'), 0) or 0)
        }
    )


def import_user(organization, row, mapping):
    """Import a user row"""
    from users.models import User
    
    email = row.get(mapping.get('email', 'email'))
    if not email:
        raise ValueError("Email is required")
    
    User.objects.update_or_create(
        email=email,
        defaults={
            'organization': organization,
            'first_name': row.get(mapping.get('first_name', 'first_name'), ''),
            'last_name': row.get(mapping.get('last_name', 'last_name'), ''),
            'is_active': True
        }
    )


def import_usage(organization, row, mapping):
    """Import usage data row"""
    from services.models import Subscription, UsageEvent
    
    subscription_name = row.get(mapping.get('subscription', 'subscription'))
    if not subscription_name:
        raise ValueError("Subscription is required")
    
    subscription = Subscription.objects.get(
        organization=organization,
        name=subscription_name
    )
    
    UsageEvent.objects.create(
        subscription=subscription,
        event_type=row.get(mapping.get('event_type', 'event_type'), 'usage'),
        timestamp=row.get(mapping.get('timestamp', 'timestamp')),
        event_data=row
    )


@shared_task
def cleanup_old_data(organization_id):
    """Clean up old data based on retention policy"""
    from users.models import Organization
    from services.models import UsageEvent, UsageMetrics
    from users.models import AuditLog
    from backups.models import DataExport
    
    try:
        org = Organization.objects.get(id=organization_id)
        
        # Usage data (default 365 days)
        usage_retention = getattr(org, 'retention_usage_days', 365)
        usage_threshold = timezone.now() - timezone.timedelta(days=usage_retention)
        
        deleted_events, _ = UsageEvent.objects.filter(
            subscription__organization=org,
            timestamp__lt=usage_threshold
        ).delete()
        
        deleted_metrics, _ = UsageMetrics.objects.filter(
            subscription__organization=org,
            period_start__lt=usage_threshold.date()
        ).delete()
        
        # Audit logs (default 365 days)
        audit_retention = getattr(org, 'retention_audit_days', 365)
        audit_threshold = timezone.now() - timezone.timedelta(days=audit_retention)
        
        deleted_audit, _ = AuditLog.objects.filter(
            organization=org,
            created_at__lt=audit_threshold
        ).delete()
        
        # Expired exports
        deleted_exports, _ = DataExport.objects.filter(
            organization=org,
            expires_at__lt=timezone.now()
        ).delete()
        
        logger.info(
            f"Cleanup for {org.name}: "
            f"events={deleted_events}, metrics={deleted_metrics}, "
            f"audit={deleted_audit}, exports={deleted_exports}"
        )
        
    except Organization.DoesNotExist:
        logger.error(f"Organization {organization_id} not found")


@shared_task
def run_scheduled_backups():
    """Check and run scheduled backups"""
    from backups.models import BackupSchedule, Backup
    
    now = timezone.now()
    
    schedules = BackupSchedule.objects.filter(
        is_active=True,
        next_run_at__lte=now
    )
    
    for schedule in schedules:
        # Create backup
        backup = Backup.objects.create(
            schedule=schedule,
            status='pending'
        )
        
        # Trigger backup task
        run_backup.delay(str(backup.id))
        
        # Calculate next run
        schedule.next_run_at = calculate_next_run(schedule)
        schedule.save()
    
    logger.info(f"Triggered {schedules.count()} scheduled backups")


def calculate_next_run(schedule):
    """Calculate next run time for schedule"""
    now = timezone.now()
    
    if schedule.frequency == 'daily':
        next_run = now.replace(
            hour=schedule.time_of_day.hour,
            minute=schedule.time_of_day.minute,
            second=0,
            microsecond=0
        )
        if next_run <= now:
            next_run += timezone.timedelta(days=1)
    
    elif schedule.frequency == 'weekly':
        days_ahead = schedule.day_of_week - now.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        next_run = now + timezone.timedelta(days=days_ahead)
        next_run = next_run.replace(
            hour=schedule.time_of_day.hour,
            minute=schedule.time_of_day.minute,
            second=0,
            microsecond=0
        )
    
    elif schedule.frequency == 'monthly':
        next_run = now.replace(
            day=schedule.day_of_month,
            hour=schedule.time_of_day.hour,
            minute=schedule.time_of_day.minute,
            second=0,
            microsecond=0
        )
        if next_run <= now:
            # Move to next month
            if now.month == 12:
                next_run = next_run.replace(year=now.year + 1, month=1)
            else:
                next_run = next_run.replace(month=now.month + 1)
    
    else:
        # Manual - no next run
        next_run = None
    
    return next_run
